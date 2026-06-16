"""把测试任务的逐条预测/评测结果导出成 xlsx(openpyxl)。两种测试任务各一个 builder。"""
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session
from app.models.prompt import PromptVersion
from app.models.llm import LlmModel
from app.models.prompt_eval import PromptEvalRun, PromptEvalItem

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _finalize(wb: Workbook) -> bytes:
    ws = wb.active
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(10, width + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- 模型测试 ----------
# 逐条预测的 key → 中文表头。不同任务类型(分类/回归对/NER/向量检索)产出的 key 不同,
# 这里做并集 + 友好表头,未知 key 原样作表头,从而兼容所有模型类型。
_PRED_HEADER = {
    "row": "序号", "input": "输入", "text": "文本", "query": "查询",
    "text_a": "文本A", "text_b": "文本B",
    "expected": "真实", "predicted": "预测",
    "expected_score": "真实分数", "predicted_score": "预测分数",
    "correct": "是否正确",
}


def _pred_cell(key, value):
    if key == "row":
        return (int(value) if value is not None else 0) + 1
    if key == "correct":
        if value is None:
            return "—"
        return "✓" if value else "✗"
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else value


def eval_predictions_xlsx(run) -> bytes:
    rows = run.predictions or []
    # 按首次出现顺序取所有 key 的并集(row 永远排第一)。
    cols: list[str] = []
    for r in rows:
        for k in r:
            if k not in cols:
                cols.append(k)
    if "row" in cols:
        cols = ["row"] + [c for c in cols if c != "row"]
    if not cols:
        cols = ["row", "input", "expected", "predicted", "correct"]
    wb = Workbook()
    ws = wb.active
    ws.title = "预测结果"
    ws.append([_PRED_HEADER.get(c, c) for c in cols])
    for r in rows:
        ws.append([_pred_cell(c, r.get(c)) for c in cols])
    return _finalize(wb)


# ---------- Prompt 评测 ----------
def _good(is_good, evaluated_at) -> str:
    if evaluated_at is None:
        return "未评"
    return "好" if is_good else "坏"


def _winner(winner_id, all_bad, evaluated_at, desc) -> str:
    if evaluated_at is None:
        return "未评"
    if all_bad:
        return "都一样坏"
    if winner_id:
        return desc.get(winner_id, "")
    return "—"


def prompt_eval_xlsx(db: Session, run: PromptEvalRun) -> bytes:
    arms = sorted(run.arms, key=lambda a: a.arm_index)
    pvs = {pv.id: pv for pv in db.execute(
        select(PromptVersion).where(PromptVersion.id.in_([a.prompt_version_id for a in arms]))
    ).scalars()} if arms else {}
    models = {m.id: m for m in db.execute(
        select(LlmModel).where(LlmModel.id.in_([a.model_id for a in arms]))
    ).scalars()} if arms else {}

    def arm_desc(a) -> str:
        pv, m = pvs.get(a.prompt_version_id), models.get(a.model_id)
        parts = []
        if pv and pv.prompt:
            parts.append(f"{pv.prompt.name} V{pv.version_no}")
        if m:
            parts.append(m.model_id)
        return " · ".join(parts) or (a.label or f"候选{a.arm_index + 1}")

    desc = {a.id: arm_desc(a) for a in arms}
    items = db.execute(select(PromptEvalItem).where(PromptEvalItem.run_id == run.id)
                       .order_by(PromptEvalItem.item_index)).scalars().all()
    keys: list[str] = []
    for it in items:
        for k in (it.inputs or {}):
            if k not in keys:
                keys.append(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "评测结果"
    single = run.eval_type == "single_prompt"
    if single:
        ws.append(["序号", *keys, "输出", "人工评估", "AI评估", "AI理由"])
        for it in items:
            out_text = it.outputs[0].output_text if it.outputs else ""
            ws.append([it.item_index + 1, *[str((it.inputs or {}).get(k, "")) for k in keys],
                       out_text, _good(it.is_good, it.evaluated_at),
                       _good(it.ai_is_good, it.ai_evaluated_at), it.ai_reasoning or ""])
    else:
        ws.append(["序号", *keys, *[f"候选:{desc[a.id]}" for a in arms],
                   "人工选择", "AI选择", "AI理由"])
        for it in items:
            by_arm = {o.arm_id: o.output_text for o in it.outputs}
            ws.append([it.item_index + 1, *[str((it.inputs or {}).get(k, "")) for k in keys],
                       *[by_arm.get(a.id, "") for a in arms],
                       _winner(it.winner_arm_id, it.all_bad, it.evaluated_at, desc),
                       _winner(it.ai_winner_arm_id, it.ai_all_bad, it.ai_evaluated_at, desc),
                       it.ai_reasoning or ""])
    return _finalize(wb)
