from fastapi.testclient import TestClient
from tests.conftest import make_user, auth_headers

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_eval_predictions(session_factory):
    from app.models.training import EvalRun
    db = session_factory()
    run = EvalRun(model_version_id=1, dataset_version_id=1, status="succeeded",
                  results={"accuracy": 0.5},
                  predictions=[{"row": 0, "input": "好评", "expected": "pos",
                                "predicted": "neg", "correct": False}])
    db.add(run); db.commit(); db.refresh(run)
    rid = run.id
    u = make_user(db, codes=("eval:read",), data_scope="all", email="ev@x.com")
    H = auth_headers(u.id); db.close()
    from app.main import app
    c = TestClient(app)
    r = c.get(f"/eval-runs/{rid}/predictions.xlsx", headers=H)
    assert r.status_code == 200 and r.headers["content-type"] == XLSX
    assert r.content[:2] == b"PK" and len(r.content) > 100  # 是个 zip(xlsx)


def test_export_eval_predictions_pair_shape(session_factory):
    # 回归对任务的逐条预测没有 correct 列,只有分数列;builder 应按 key 动态出表头。
    from app.services import export_service
    from app.models.training import EvalRun
    run = EvalRun(model_version_id=1, dataset_version_id=1, status="succeeded", results={},
                  predictions=[{"row": 0, "text_a": "猫", "text_b": "猫咪",
                                "expected_score": 1.0, "predicted_score": 0.93}])
    data = export_service.eval_predictions_xlsx(run)
    assert data[:2] == b"PK" and len(data) > 100

    import io
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(data)).active
    headers = [c.value for c in ws[1]]
    assert headers == ["序号", "文本A", "文本B", "真实分数", "预测分数"]
    assert [c.value for c in ws[2]] == [1, "猫", "猫咪", 1.0, 0.93]


def test_export_eval_no_predictions_409(session_factory):
    from app.models.training import EvalRun
    db = session_factory()
    run = EvalRun(model_version_id=1, dataset_version_id=1, status="succeeded",
                  results={"accuracy": 1.0}, predictions=[])
    db.add(run); db.commit(); db.refresh(run)
    rid = run.id
    u = make_user(db, codes=("eval:read",), data_scope="all", email="ev2@x.com")
    H = auth_headers(u.id); db.close()
    from app.main import app
    c = TestClient(app)
    assert c.get(f"/eval-runs/{rid}/predictions.xlsx", headers=H).status_code == 409


def test_export_eval_requires_perm(session_factory):
    db = session_factory()
    u = make_user(db, codes=("prompteval:read",), data_scope="all", email="np2@x.com")
    H = auth_headers(u.id); db.close()
    from app.main import app
    c = TestClient(app)
    assert c.get("/eval-runs/1/predictions.xlsx", headers=H).status_code == 403


def test_export_prompt_eval_results(session_factory):
    from app.models.prompt_eval import PromptEvalRun, PromptEvalArm, PromptEvalItem, PromptEvalOutput
    db = session_factory()
    run = PromptEvalRun(name="r", eval_type="multi_model", prompt_version_ids=[1],
                        model_ids=[1, 2], dataset_version_ids=[1])
    run.arms.append(PromptEvalArm(arm_index=0, prompt_version_id=1, model_id=1, label="A"))
    run.arms.append(PromptEvalArm(arm_index=1, prompt_version_id=1, model_id=2, label="B"))
    db.add(run); db.commit(); db.refresh(run)
    it = PromptEvalItem(run_id=run.id, item_index=0, dataset_version_id=1, row_index=0,
                        inputs={"问题": "你好"}, winner_arm_id=run.arms[0].id, evaluated_at=None)
    it.outputs.append(PromptEvalOutput(arm_id=run.arms[0].id, output_text="答A", status="done"))
    it.outputs.append(PromptEvalOutput(arm_id=run.arms[1].id, output_text="答B", status="done"))
    db.add(it); db.commit()
    rid = run.id
    u = make_user(db, codes=("prompteval:read",), data_scope="all", email="pe@x.com")
    H = auth_headers(u.id); db.close()
    from app.main import app
    c = TestClient(app)
    r = c.get(f"/prompt-evals/{rid}/results.xlsx", headers=H)
    assert r.status_code == 200 and r.headers["content-type"] == XLSX
    assert r.content[:2] == b"PK"
